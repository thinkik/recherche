#!/usr/bin/env python3
"""Fügt eine Rückmeldung in die erste freie Zeile der Excel-Datei ein."""

from pathlib import Path

from openpyxl import load_workbook

FILE_PATH = Path(__file__).resolve().parent / "rueckmeldungen.xlsx"


def first_free_row(worksheet, columns=(1, 2, 3)) -> int:
    for row in range(1, worksheet.max_row + 1):
        if all(worksheet.cell(row=row, column=col).value in (None, "") for col in columns):
            return row
    return worksheet.max_row + 1


def main() -> None:
    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active

    target_row = first_free_row(worksheet)
    worksheet.cell(row=target_row, column=1, value="Lukas")
    worksheet.cell(row=target_row, column=2, value="Hubolt")
    worksheet.cell(row=target_row, column=3, value="Sehr gute Einführung")

    workbook.save(FILE_PATH)


if __name__ == "__main__":
    main()
